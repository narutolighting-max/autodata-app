"""
AutoData Technologies — Servidor de Producción v3.0.0
Backend Flask con API REST + Frontend integrado
© 2025 AutoData Technologies — autodata.com
"""

import os
import json
import sqlite3
import hashlib
import secrets
import re
import time
import random
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ═══════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════
VERSION = "3.1.0"
SECRET_KEY = os.environ.get("AUTODATA_SECRET", secrets.token_hex(32))
PORT = int(os.environ.get("PORT", 5000))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Railway Volume: si RAILWAY_VOLUME_MOUNT_PATH está definido, usar ese directorio
# para DB, data y uploads (persiste entre deploys)
PERSIST_DIR = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "")
if PERSIST_DIR:
    DB_PATH = os.path.join(PERSIST_DIR, "autodata.db")
    DATA_DIR = os.path.join(PERSIST_DIR, "data")
    UPLOAD_DIR = os.path.join(PERSIST_DIR, "uploads")
else:
    DB_PATH = os.environ.get("DATABASE_URL", "autodata.db")
    DATA_DIR = os.path.join(BASE_DIR, "data")
    UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# SendGrid para verificación por email
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
SENDGRID_FROM_EMAIL = os.environ.get("SENDGRID_FROM_EMAIL", "noreply@dataintelligence.com")

# Caché de códigos de verificación: {email: {"code": "123456", "expira": datetime}}
_verification_codes = {}

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["SECRET_KEY"] = SECRET_KEY
CORS(app, resources={r"/ad-api/*": {"origins": "*"}})


# ═══════════════════════════════════════════════
# BASE DE DATOS
# ═══════════════════════════════════════════════

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        nombre TEXT,
        rol TEXT NOT NULL DEFAULT 'operator',
        empresa TEXT NOT NULL DEFAULT 'default',
        empresa_id TEXT NOT NULL DEFAULT 'default',
        activo INTEGER DEFAULT 1,
        ultimo_login TEXT,
        creado TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS sesiones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER,
        token TEXT UNIQUE NOT NULL,
        creado TEXT DEFAULT CURRENT_TIMESTAMP,
        expira TEXT NOT NULL,
        activa INTEGER DEFAULT 1
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS historial_procesamiento (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        empresa_id TEXT NOT NULL,
        usuario_email TEXT,
        fecha TEXT DEFAULT CURRENT_TIMESTAMP,
        archivos_total INTEGER DEFAULT 0,
        facturas_nuevas INTEGER DEFAULT 0,
        duplicadas INTEGER DEFAULT 0,
        otros_docs INTEGER DEFAULT 0,
        detalle TEXT
    )""")

    conn.commit()

    # Crear usuarios iniciales si no existen
    if c.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
        c.execute("""INSERT INTO usuarios (email, password_hash, nombre, rol, empresa, empresa_id)
            VALUES (?, ?, ?, ?, ?, ?)""",
            ('demo@dataintelligence.com', _hash_password('AutoData2025'),
             'Guillermo', 'admin', 'Dataintelligence.com', 'dataintelligence'))

        c.execute("""INSERT INTO usuarios (email, password_hash, nombre, rol, empresa, empresa_id)
            VALUES (?, ?, ?, ?, ?, ?)""",
            ('operadora@dataintelligence.com', _hash_password('Operadora2025'),
             'Operadora', 'operator', 'Dataintelligence.com', 'dataintelligence'))
        conn.commit()

    conn.close()


def _hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# ═══════════════════════════════════════════════
# AUTENTICACIÓN
# ═══════════════════════════════════════════════

def generar_token(usuario_id):
    token = secrets.token_urlsafe(32)
    expira = (datetime.utcnow() + timedelta(hours=8)).isoformat()
    conn = get_db()
    conn.execute("INSERT INTO sesiones (usuario_id, token, expira) VALUES (?,?,?)",
                 (usuario_id, token, expira))
    conn.commit()
    conn.close()
    return token, expira


def verificar_token(token):
    if not token:
        return None
    conn = get_db()
    row = conn.execute("""
        SELECT u.id, u.email, u.nombre, u.rol, u.empresa, u.empresa_id
        FROM sesiones s JOIN usuarios u ON s.usuario_id = u.id
        WHERE s.token=? AND s.activa=1 AND s.expira > ?
    """, (token, datetime.utcnow().isoformat())).fetchone()
    conn.close()
    return dict(row) if row else None


def requiere_auth(f):
    @wraps(f)
    def decorado(*args, **kwargs):
        token = request.headers.get("X-AutoData-Token") or request.args.get("token")
        usuario = verificar_token(token)
        if not usuario:
            return jsonify({"error": True, "mensaje": "Sesión inválida o expirada."}), 401
        request.usuario = usuario
        return f(*args, **kwargs)
    return decorado


def get_empresa_excel(empresa_id):
    """Retorna la ruta del Excel para una empresa específica."""
    return os.path.join(DATA_DIR, f"repositorio_{empresa_id}.xlsx")


# ═══════════════════════════════════════════════
# HEADERS & ROUTES
# ═══════════════════════════════════════════════

@app.after_request
def branded_headers(response):
    response.headers["Server"] = f"AutoData/{VERSION}"
    response.headers.pop("X-Powered-By", None)
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-AutoData-Token"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.before_request
def handle_options():
    if request.method == "OPTIONS":
        return "", 200


@app.route("/")
def frontend():
    return send_from_directory("templates", "index.html")


# ═══════════════════════════════════════════════
# API — AUTH
# ═══════════════════════════════════════════════

@app.route("/ad-api/auth/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")

    if not email or not password:
        return jsonify({"error": True, "mensaje": "Email y contraseña son requeridos."}), 400

    conn = get_db()
    usuario = conn.execute("SELECT * FROM usuarios WHERE email=? AND activo=1", (email,)).fetchone()
    conn.close()

    if not usuario or usuario["password_hash"] != _hash_password(password):
        time.sleep(0.5)
        return jsonify({"error": True, "mensaje": "Credenciales incorrectas."}), 401

    token, expira = generar_token(usuario["id"])

    conn = get_db()
    conn.execute("UPDATE usuarios SET ultimo_login=? WHERE id=?",
                 (datetime.utcnow().isoformat(), usuario["id"]))
    conn.commit()
    conn.close()

    return jsonify({
        "error": False,
        "token": token,
        "expira": expira,
        "usuario": {
            "email": usuario["email"],
            "nombre": usuario["nombre"],
            "rol": usuario["rol"],
            "empresa": usuario["empresa"],
            "empresa_id": usuario["empresa_id"]
        },
        "mensaje": f"Bienvenido, {usuario['nombre']}."
    })


@app.route("/ad-api/auth/logout", methods=["POST"])
@requiere_auth
def logout():
    token = request.headers.get("X-AutoData-Token")
    conn = get_db()
    conn.execute("UPDATE sesiones SET activa=0 WHERE token=?", (token,))
    conn.commit()
    conn.close()
    return jsonify({"mensaje": "Sesión cerrada."})


def _send_email_sendgrid(to_email, subject, html_content):
    """Enviar email usando SendGrid Web API v3 (sin dependencias externas)."""
    if not SENDGRID_API_KEY:
        return False, "SendGrid no configurado."
    payload = json.dumps({
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": SENDGRID_FROM_EMAIL, "name": "AutoData Technologies"},
        "subject": subject,
        "content": [{"type": "text/html", "value": html_content}]
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=payload,
        headers={
            "Authorization": f"Bearer {SENDGRID_API_KEY}",
            "Content-Type": "application/json"
        },
        method="POST"
    )
    try:
        resp = urllib.request.urlopen(req, timeout=10)
        return True, "OK"
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return False, f"SendGrid HTTP {e.code}: {body}"
    except Exception as e:
        return False, str(e)


def _limpiar_codigos_expirados():
    """Eliminar códigos vencidos del caché."""
    ahora = datetime.utcnow()
    expirados = [k for k, v in _verification_codes.items() if v["expira"] < ahora]
    for k in expirados:
        del _verification_codes[k]


@app.route("/ad-api/auth/request-code", methods=["POST"])
def request_verification_code():
    """Envía un código de 6 dígitos al email del usuario para cambio de contraseña."""
    data = request.get_json() or {}
    email = data.get("email", "").strip().lower()

    if not email:
        return jsonify({"error": "Se requiere el email."}), 400

    # Verificar que el email existe
    conn = get_db()
    user = conn.execute("SELECT id, nombre FROM usuarios WHERE email=? AND activo=1",
                        (email,)).fetchone()
    conn.close()

    if not user:
        return jsonify({"error": "No se encontró una cuenta con ese email."}), 404

    # Rate limiting: no enviar si ya hay un código vigente hace menos de 60 seg
    _limpiar_codigos_expirados()
    if email in _verification_codes:
        creado = _verification_codes[email].get("creado")
        if creado and (datetime.utcnow() - creado).total_seconds() < 60:
            return jsonify({"error": "Ya se envió un código. Esperá 60 segundos."}), 429

    # Generar código de 6 dígitos
    code = str(random.randint(100000, 999999))
    _verification_codes[email] = {
        "code": code,
        "expira": datetime.utcnow() + timedelta(minutes=10),
        "creado": datetime.utcnow(),
        "intentos": 0
    }

    # Enviar email
    nombre = user["nombre"] or email
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;padding:20px">
        <h2 style="color:#1a2233">AutoData Technologies</h2>
        <p>Hola <strong>{nombre}</strong>,</p>
        <p>Tu código de verificación para cambiar la contraseña es:</p>
        <div style="background:#f0f4ff;border-radius:12px;padding:20px;text-align:center;margin:20px 0">
            <span style="font-size:32px;font-weight:700;letter-spacing:8px;color:#2563eb">{code}</span>
        </div>
        <p style="color:#666;font-size:14px">Este código expira en <strong>10 minutos</strong>.</p>
        <p style="color:#666;font-size:14px">Si no solicitaste este cambio, ignorá este email.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="color:#999;font-size:12px">AutoData Technologies — dataintelligence.com</p>
    </div>
    """
    ok, msg = _send_email_sendgrid(email, "Tu código de verificación — AutoData", html)

    if not ok:
        del _verification_codes[email]
        return jsonify({"error": f"No se pudo enviar el email. {msg}"}), 500

    return jsonify({"mensaje": "Código enviado a tu email.", "email_enviado": True})


@app.route("/ad-api/auth/change-password", methods=["POST"])
def change_password():
    data = request.get_json() or {}
    email = data.get("email", "").strip().lower()
    code = data.get("code", "").strip()
    current_pw = data.get("current_password", "").strip()
    new_pw = data.get("new_password", "").strip()

    if not email or not code or not current_pw or not new_pw:
        return jsonify({"error": "Completá todos los campos."}), 400
    if len(new_pw) < 8:
        return jsonify({"error": "La nueva contraseña debe tener al menos 8 caracteres."}), 400

    # Verificar código
    _limpiar_codigos_expirados()
    stored = _verification_codes.get(email)
    if not stored:
        return jsonify({"error": "No hay un código vigente. Solicitá uno nuevo."}), 400

    stored["intentos"] = stored.get("intentos", 0) + 1
    if stored["intentos"] > 5:
        del _verification_codes[email]
        return jsonify({"error": "Demasiados intentos. Solicitá un nuevo código."}), 429

    if stored["code"] != code:
        return jsonify({"error": f"Código incorrecto. Te quedan {5 - stored['intentos']} intentos."}), 403

    # Código correcto — verificar contraseña actual
    conn = get_db()
    user = conn.execute("SELECT id, password_hash FROM usuarios WHERE email=?",
                        (email,)).fetchone()

    if not user or user["password_hash"] != _hash_password(current_pw):
        conn.close()
        return jsonify({"error": "Contraseña actual incorrecta."}), 403

    conn.execute("UPDATE usuarios SET password_hash=? WHERE id=?",
                 (_hash_password(new_pw), user["id"]))
    conn.execute("UPDATE sesiones SET activa=0 WHERE usuario_id=?", (user["id"],))
    conn.commit()
    conn.close()

    # Limpiar código usado
    del _verification_codes[email]
    return jsonify({"mensaje": "Contraseña actualizada. Iniciá sesión con tu nueva clave."})


# ═══════════════════════════════════════════════
# EXCEL — CREACIÓN Y GESTIÓN (MULTI-TENANT)
# ═══════════════════════════════════════════════

EXCEL_HEADERS = [
    'Tipo Comp.', 'Punto Vta.', 'Nro. Comp.', 'Fecha Emisión',
    'CUIT Emisor', 'Razón Social Emisor', 'Domicilio Emisor', 'Cond. IVA Emisor',
    'CUIT Cliente', 'Razón Social Cliente', 'Domicilio Cliente', 'Cond. IVA Cliente',
    'Cond. Venta', 'Per. Desde', 'Per. Hasta', 'Vto. Pago',
    'Producto / Servicio', 'Importe Total ($)', 'CAE N°', 'Vto. CAE',
    'Categoría', 'Centro de Costo', 'Importe USD', 'Tipo Cambio', 'Alertas', 'Archivo PDF'
]

HEADER_GROUPS = [
    ('A3', 'C3', 'COMPROBANTE'), ('D3', 'H3', 'EMISOR'),
    ('I3', 'M3', 'CLIENTE'), ('N3', 'P3', 'PERÍODO / VTO.'),
    ('Q3', 'Q3', 'DETALLE'), ('R3', 'T3', 'IMPORTES / CAE'),
    ('U3', 'V3', 'CLASIFICACIÓN'), ('W3', 'X3', 'MULTIMONEDA'), ('Y3', 'Z3', 'AUDITORÍA'),
]

COL_WIDTHS = {
    'A': 12, 'B': 10, 'C': 14, 'D': 14, 'E': 18, 'F': 28,
    'G': 36, 'H': 18, 'I': 18, 'J': 32, 'K': 40, 'L': 20,
    'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 24, 'R': 18,
    'S': 20, 'T': 12, 'U': 16, 'V': 20, 'W': 16, 'X': 14, 'Y': 40, 'Z': 44
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") if EXCEL_SUPPORT else None
DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") if EXCEL_SUPPORT else None
MED_BLUE = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid") if EXCEL_SUPPORT else None
WHITE_BOLD_13 = Font(name="Arial", size=13, bold=True, color="FFFFFF") if EXCEL_SUPPORT else None
WHITE_BOLD_9 = Font(name="Arial", size=9, bold=True, color="FFFFFF") if EXCEL_SUPPORT else None
DATA_FONT = Font(name="Arial", size=9) if EXCEL_SUPPORT else None
CENTER = Alignment(horizontal="center", vertical="center") if EXCEL_SUPPORT else None
LEFT = Alignment(horizontal="left", vertical="center") if EXCEL_SUPPORT else None


def get_or_create_excel(empresa_id, empresa_nombre=""):
    if not EXCEL_SUPPORT:
        return None
    excel_path = get_empresa_excel(empresa_id)

    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            if ws.cell(row=4, column=1).value == 'Tipo Comp.' and ws.max_column >= 26:
                return wb
            wb.close()
            os.remove(excel_path)
        except Exception:
            try:
                os.remove(excel_path)
            except Exception:
                pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Repositorio Facturas"

    ws.merge_cells('A1:Z1')
    title_cell = ws['A1']
    title_cell.value = f"REPOSITORIO DE FACTURAS — {empresa_nombre.upper()}" if empresa_nombre else "REPOSITORIO DE FACTURAS"
    title_cell.font = WHITE_BOLD_13
    title_cell.fill = DARK_BLUE
    title_cell.alignment = CENTER

    for start, end, label in HEADER_GROUPS:
        if start != end:
            ws.merge_cells(f'{start}:{end}')
        cell = ws[start]
        cell.value = label
        cell.font = WHITE_BOLD_9
        cell.fill = DARK_BLUE
        cell.alignment = CENTER

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = WHITE_BOLD_9
        cell.fill = MED_BLUE
        cell.alignment = CENTER

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A5'

    # Otros Documentos sheet
    ws2 = wb.create_sheet("Otros Documentos")
    ws2.merge_cells('A1:E1')
    ws2['A1'].value = "DOCUMENTOS NO RELACIONADOS A FACTURACIÓN"
    ws2['A1'].font = WHITE_BOLD_13
    ws2['A1'].fill = DARK_BLUE
    ws2['A1'].alignment = CENTER
    ws2.merge_cells('A2:E2')
    ws2['A2'].value = "Archivos que NO son comprobantes de venta"
    ws2['A2'].font = Font(name="Arial", size=9, italic=True, color="666666")
    for col_idx, h in enumerate(['Nombre del Archivo', 'Tipo', 'Observación', 'Fecha Detección', 'Acción Sugerida'], 1):
        cell = ws2.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = WHITE_BOLD_9
        cell.fill = MED_BLUE
        cell.alignment = CENTER
    ws2.column_dimensions['A'].width = 42
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 60
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 36
    ws2.freeze_panes = 'A5'

    # Resumen sheet
    ws3 = wb.create_sheet("Resumen")
    ws3.merge_cells('A1:B1')
    ws3['A1'].value = "RESUMEN EJECUTIVO"
    ws3['A1'].font = WHITE_BOLD_13
    ws3['A1'].fill = DARK_BLUE
    ws3['A1'].alignment = CENTER
    ws3['A2'].value = "Concepto"
    ws3['A2'].font = WHITE_BOLD_9
    ws3['A2'].fill = MED_BLUE
    ws3['B2'].value = "Valor"
    ws3['B2'].font = WHITE_BOLD_9
    ws3['B2'].fill = MED_BLUE
    ws3['A3'].value = "Cantidad de Facturas"
    ws3['B3'] = "=COUNTA('Repositorio Facturas'!C5:C9999)"
    ws3['A4'].value = "Importe total facturado (ARS)"
    ws3['B4'] = "=SUM('Repositorio Facturas'!R5:R9999)"
    ws3['A5'].value = "Importe total facturado (USD)"
    ws3['B5'] = "=SUM('Repositorio Facturas'!W5:W9999)"
    ws3['A6'].value = "Facturas con alertas"
    ws3['B6'] = "=COUNTIF('Repositorio Facturas'!Y5:Y9999,\"<>OK\")"
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 24

    wb.save(excel_path)
    return wb


def get_existing_invoices(ws):
    existing = set()
    if ws:
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row and len(row) > 2 and row[2]:
                existing.add(str(row[2]).strip())
    return existing


def get_existing_otros(ws):
    existing = set()
    if ws:
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row and row[0]:
                existing.add(str(row[0]).strip())
    return existing


# ═══════════════════════════════════════════════
# PDF EXTRACTION
# ═══════════════════════════════════════════════

def format_cuit(raw):
    raw = raw.strip()
    if '-' in raw:
        return raw
    if len(raw) == 11:
        return f"{raw[:2]}-{raw[2:10]}-{raw[10]}"
    return raw


def clean_field(value, stop_words=None):
    if not value:
        return ""
    if stop_words:
        for sw in stop_words:
            idx = value.lower().find(sw.lower())
            if idx > 0:
                value = value[:idx]
    return value.strip()


# ═══════════════════════════════════════════════
# DETECCIÓN DE ANOMALÍAS (AUDITOR AUTOMÁTICO)
# ═══════════════════════════════════════════════

def validar_cuit(cuit_str):
    """Valida dígito verificador de CUIT argentino."""
    digits = cuit_str.replace('-', '')
    if len(digits) != 11 or not digits.isdigit():
        return False
    mult = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    total = sum(int(d) * m for d, m in zip(digits[:10], mult))
    check = 11 - (total % 11)
    if check == 11:
        check = 0
    elif check == 10:
        check = 9
    return check == int(digits[10])


def detectar_anomalias(data):
    """Analiza una factura y devuelve lista de alertas."""
    alertas = []

    # 1. CUIT inválido
    cuit_e = data.get('cuit_emisor', '')
    if cuit_e and not validar_cuit(cuit_e):
        alertas.append("CUIT Emisor inválido")
    cuit_c = data.get('cuit_cliente', '')
    if cuit_c and not validar_cuit(cuit_c):
        alertas.append("CUIT Cliente inválido")

    # 2. CUIT emisor = CUIT cliente (auto-factura)
    if cuit_e and cuit_c and cuit_e.replace('-', '') == cuit_c.replace('-', ''):
        alertas.append("CUIT Emisor = Cliente (auto-factura)")

    # 3. Importe cero o negativo
    importe = data.get('importe', 0)
    if not importe or importe <= 0:
        alertas.append("Importe $0 o negativo")

    # 4. Importe muy alto (> $50M ARS)
    if importe and importe > 50000000:
        alertas.append(f"Importe inusual: ${importe:,.0f}")

    # 5. CAE vencido
    vto_cae = data.get('vto_cae', '')
    if vto_cae:
        try:
            for fmt in ('%d/%m/%Y', '%d-%m-%Y'):
                try:
                    vto_date = datetime.strptime(vto_cae, fmt)
                    if vto_date < datetime.now():
                        alertas.append("CAE vencido")
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # 6. Fecha futura
    fecha = data.get('fecha', '')
    if fecha:
        try:
            for fmt in ('%d/%m/%Y', '%d-%m-%Y'):
                try:
                    f_date = datetime.strptime(fecha, fmt)
                    if f_date > datetime.now() + timedelta(days=1):
                        alertas.append("Fecha de emisión futura")
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # 7. Campos críticos faltantes
    if not data.get('cae'):
        alertas.append("Sin CAE")
    if not data.get('cuit_emisor'):
        alertas.append("Sin CUIT Emisor")
    if not data.get('razon_social_emisor'):
        alertas.append("Sin Razón Social Emisor")

    # 8. Condición IVA inconsistente
    tipo = data.get('tipo', '').lower()
    cond_cliente = data.get('cond_iva_cliente', '').lower()
    if 'factura a' in tipo and 'monotributo' in cond_cliente:
        alertas.append("Factura A a Monotributista (revisar)")

    return alertas


# ═══════════════════════════════════════════════
# CATEGORIZACIÓN INTELIGENTE
# ═══════════════════════════════════════════════

CATEGORIAS_REGLAS = [
    # (palabras clave en producto/razón social, categoría, centro de costo)
    (['honorarios', 'servicio profesional', 'consultor', 'asesor'], 'Servicio Profesional', 'Servicios'),
    (['hosting', 'dominio', 'cloud', 'software', 'licencia', 'saas', 'aws', 'google cloud'], 'Tecnología', 'IT'),
    (['alquiler', 'expensas', 'inmobiliario'], 'Alquiler / Inmueble', 'Inmueble'),
    (['seguro', 'póliza', 'cobertura'], 'Seguros', 'Seguros'),
    (['electricidad', 'gas', 'agua', 'internet', 'teléfono', 'celular', 'edenor', 'edesur', 'metrogas', 'telecom', 'movistar', 'claro', 'personal'], 'Servicios Públicos', 'Servicios'),
    (['combustible', 'nafta', 'gasoil', 'ypf', 'shell', 'axion'], 'Combustible', 'Transporte'),
    (['supermercado', 'alimento', 'comida', 'catering', 'restaurant'], 'Alimentación', 'Gastos Generales'),
    (['publicidad', 'marketing', 'google ads', 'facebook', 'meta', 'campaña'], 'Marketing / Publicidad', 'Marketing'),
    (['impuesto', 'iibb', 'ingresos brutos', 'afip', 'arba', 'agip', 'tasa', 'contribución'], 'Impuestos / Tasas', 'Impuestos'),
    (['contable', 'contador', 'auditor', 'estudio'], 'Honorarios Contables', 'Servicios'),
    (['transporte', 'flete', 'envío', 'logística', 'correo', 'oca', 'andreani'], 'Transporte / Logística', 'Logística'),
    (['mueble', 'equipo', 'computadora', 'notebook', 'monitor', 'impresora'], 'Equipamiento', 'Activos Fijos'),
    (['papelería', 'librería', 'insumo', 'oficina', 'toner', 'resma'], 'Insumos Oficina', 'Gastos Generales'),
]


def categorizar_factura(data):
    """Asigna categoría y centro de costo basado en reglas."""
    tipo = data.get('tipo', '').lower()

    # Primero determinar si es Venta o Gasto por el tipo de comprobante
    es_nota_credito = 'nota de cr' in tipo
    es_nota_debito = 'nota de d' in tipo

    # Buscar match por producto + razón social
    texto_busqueda = (
        (data.get('producto', '') + ' ' +
         data.get('razon_social_emisor', '') + ' ' +
         data.get('razon_social_cliente', '')).lower()
    )

    for keywords, categoria, centro in CATEGORIAS_REGLAS:
        for kw in keywords:
            if kw in texto_busqueda:
                if es_nota_credito:
                    return f"NC - {categoria}", centro
                elif es_nota_debito:
                    return f"ND - {categoria}", centro
                return categoria, centro

    # Default
    if 'factura' in tipo:
        return 'Gasto General', 'Sin Clasificar'
    elif es_nota_credito:
        return 'Nota de Crédito', 'Sin Clasificar'
    elif es_nota_debito:
        return 'Nota de Débito', 'Sin Clasificar'
    elif 'recibo' in tipo:
        return 'Recibo', 'Sin Clasificar'

    return 'Sin Categoría', 'Sin Clasificar'


# ═══════════════════════════════════════════════
# SOPORTE MULTIMONEDA (ARS + USD)
# ═══════════════════════════════════════════════

_cached_usd_rate = {"rate": None, "date": None}


def get_usd_rate():
    """Obtiene tipo de cambio oficial USD/ARS del día. Cache por fecha."""
    today = datetime.now().strftime("%Y-%m-%d")
    if _cached_usd_rate["date"] == today and _cached_usd_rate["rate"]:
        return _cached_usd_rate["rate"]
    try:
        import urllib.request
        import json as _json
        # Bluelytics API — tipo de cambio oficial + blue Argentina
        req = urllib.request.Request("https://api.bluelytics.com.ar/v2/latest",
                                     headers={"User-Agent": "AutoData/3.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = _json.loads(resp.read())
            # Usar cotización oficial venta
            rate = data.get("oficial", {}).get("value_sell", 0)
            if rate and rate > 0:
                _cached_usd_rate["rate"] = round(rate, 2)
                _cached_usd_rate["date"] = today
                return _cached_usd_rate["rate"]
    except Exception:
        pass
    try:
        import urllib.request
        import json as _json
        req = urllib.request.Request("https://dolarapi.com/v1/dolares/oficial",
                                     headers={"User-Agent": "AutoData/3.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = _json.loads(resp.read())
            rate = data.get("venta", 0)
            if rate and rate > 0:
                _cached_usd_rate["rate"] = round(rate, 2)
                _cached_usd_rate["date"] = today
                return _cached_usd_rate["rate"]
    except Exception:
        pass
    # Fallback
    return _cached_usd_rate.get("rate") or 0


def convertir_a_usd(importe_ars, tc=None):
    """Convierte ARS a USD usando tipo de cambio del día."""
    if not tc:
        tc = get_usd_rate()
    if tc and tc > 0 and importe_ars:
        return round(importe_ars / tc, 2)
    return 0


# ═══════════════════════════════════════════════
# PDF EXTRACTION
# ═══════════════════════════════════════════════

def extract_invoice_data(pdf_path):
    if not PDF_SUPPORT:
        return None
    try:
        data = {}
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                return None
            text = pdf.pages[0].extract_text()
            if not text:
                return None

            tipo_match = re.search(r'(FACTURA|RECIBO|NOTA DE CR[ÉE]DITO|NOTA DE D[ÉE]BITO)', text, re.IGNORECASE)
            letra_match = re.search(r'^([ABC])\s*$', text, re.MULTILINE)
            if tipo_match:
                tipo_name = tipo_match.group(1).strip().title()
                tipo_letra = letra_match.group(1) if letra_match else ""
                data['tipo'] = f"{tipo_name} {tipo_letra}".strip()
            else:
                data['tipo'] = ""

            pv_match = re.search(r'Punto de Venta:\s*(\d+)\s*Comp\.?\s*Nro:?\s*(\d+)', text, re.IGNORECASE)
            if pv_match:
                data['punto_vta'] = pv_match.group(1).zfill(5)
                data['nro_comp'] = pv_match.group(2).zfill(8)
            else:
                nro_match = re.search(r'(?:Comp\.?\s*Nro|N[°º]|Nro\.?|N[úu]mero)[\s.:]*(\d[\d-]*\d)', text, re.IGNORECASE)
                data['nro_comp'] = nro_match.group(1).strip() if nro_match else ""
                data['punto_vta'] = ""

            fecha_em = re.search(r'Fecha\s*de\s*Emisi[óo]n[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            if not fecha_em:
                fecha_em = re.search(r'Fecha[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            data['fecha'] = fecha_em.group(1) if fecha_em else ""

            cuit_all = re.findall(r'CUIT[:\s]*(\d{11}|\d{2}-\d{8}-\d)', text)
            data['cuit_emisor'] = format_cuit(cuit_all[0]) if len(cuit_all) > 0 else ""
            data['cuit_cliente'] = format_cuit(cuit_all[1]) if len(cuit_all) > 1 else ""

            razones = re.findall(r'(?:Raz[óo]n\s*Social|Apellido\s*y\s*Nombre\s*/?\s*Raz[óo]n\s*Social)[:\s]*([^\n]+)', text, re.IGNORECASE)
            data['razon_social_emisor'] = clean_field(razones[0] if len(razones) > 0 else "", ['Fecha de', 'CUIT', 'Domicilio'])[:80]
            data['razon_social_cliente'] = clean_field(razones[1] if len(razones) > 1 else "", ['Fecha de', 'CUIT', 'Domicilio'])[:80]

            domicilios = re.findall(r'Domicilio\s*(?:Comercial)?[:\s]*([^\n]+)', text, re.IGNORECASE)
            data['domicilio_emisor'] = clean_field(domicilios[0] if len(domicilios) > 0 else "", ['CUIT', 'Ingresos', 'Condición'])[:100]
            if len(domicilios) > 1:
                data['domicilio_cliente'] = clean_field(domicilios[1], ['CUIT', 'Ingresos'])[:100]
            else:
                dom2 = re.search(r'Domicilio[:\s]*([^\n]+)', text[text.find('Apellido'):] if 'Apellido' in text else '', re.IGNORECASE)
                data['domicilio_cliente'] = dom2.group(1).strip()[:100] if dom2 else ""

            iva_all = re.findall(r'Condici[óo]n\s*frente\s*al\s*IVA[:\s]*([^\n]+)', text, re.IGNORECASE)
            data['cond_iva_emisor'] = clean_field(iva_all[0] if len(iva_all) > 0 else "", ['Fecha de', 'Domicilio', 'Inicio'])[:40]
            data['cond_iva_cliente'] = clean_field(iva_all[1] if len(iva_all) > 1 else "", ['Fecha de', 'Domicilio', 'Inicio'])[:40]
            for key in ['cond_iva_emisor', 'cond_iva_cliente']:
                v = data[key].lower()
                if 'monotributo' in v:
                    data[key] = 'Resp. Monotributo'
                elif 'responsable inscripto' in v:
                    data[key] = 'IVA Resp. Inscripto'
                elif 'exento' in v:
                    data[key] = 'IVA Exento'

            cond_vta = re.search(r'Condici[óo]n\s*de\s*venta[:\s]*([^\n]+)', text, re.IGNORECASE)
            data['cond_venta'] = cond_vta.group(1).strip()[:30] if cond_vta else ""

            per_desde = re.search(r'(?:Per[íi]odo\s*Facturado\s*)?Desde[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            per_hasta = re.search(r'Hasta[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            data['per_desde'] = per_desde.group(1) if per_desde else ""
            data['per_hasta'] = per_hasta.group(1) if per_hasta else ""

            vto_pago = re.search(r'(?:Fecha\s*de\s*)?Vto\.?\s*(?:para\s*el\s*)?(?:pago|Pago)[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            data['vto_pago'] = vto_pago.group(1) if vto_pago else ""

            lines = text.split('\n')
            data['producto'] = ""
            for i, line in enumerate(lines):
                if 'Producto / Servicio' in line or 'Producto/Servicio' in line:
                    if i + 1 < len(lines):
                        item = re.match(r'\d+\s+(.+?)\s+\d+[.,]', lines[i + 1].strip())
                        if item:
                            data['producto'] = item.group(1).strip()[:100]
                    break

            imp_match = re.search(r'Importe\s*Total[:\s$]*\$?\s*([\d.,]+)', text, re.IGNORECASE)
            if imp_match:
                monto_str = imp_match.group(1).replace('.', '').replace(',', '.')
                try:
                    data['importe'] = float(monto_str)
                except Exception:
                    data['importe'] = 0.0
            else:
                data['importe'] = 0.0

            cae_match = re.search(r'CAE\s*N?[°º]?\s*:?\s*(\d{10,14})', text, re.IGNORECASE)
            data['cae'] = cae_match.group(1) if cae_match else ""

            vto_cae = re.search(r'(?:Fecha\s*de\s*)?Vto\.?\s*(?:de\s*)?CAE[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
            data['vto_cae'] = vto_cae.group(1) if vto_cae else ""

            return data
    except Exception as e:
        print(f"Error extrayendo PDF {pdf_path}: {e}")
        return None


ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") if EXCEL_SUPPORT else None
RED_FONT = Font(name="Arial", size=9, color="CC0000") if EXCEL_SUPPORT else None


def add_invoice_row(ws, row_num, data, filename):
    # Campos base (cols 1-20)
    fields = [
        ('tipo', 1), ('punto_vta', 2), ('nro_comp', 3), ('fecha', 4),
        ('cuit_emisor', 5), ('razon_social_emisor', 6), ('domicilio_emisor', 7),
        ('cond_iva_emisor', 8), ('cuit_cliente', 9), ('razon_social_cliente', 10),
        ('domicilio_cliente', 11), ('cond_iva_cliente', 12), ('cond_venta', 13),
        ('per_desde', 14), ('per_hasta', 15), ('vto_pago', 16),
        ('producto', 17), ('importe', 18), ('cae', 19), ('vto_cae', 20)
    ]
    missing = []
    for key, col in fields:
        cell = ws.cell(row=row_num, column=col)
        val = data.get(key, "")
        cell.value = val
        cell.font = DATA_FONT
        cell.alignment = LEFT if col > 4 else CENTER
        if not val and val != 0:
            cell.fill = YELLOW_FILL
            missing.append(EXCEL_HEADERS[col - 1])

    # Categorización Inteligente (cols 21-22)
    categoria, centro = categorizar_factura(data)
    ws.cell(row=row_num, column=21).value = categoria
    ws.cell(row=row_num, column=21).font = DATA_FONT
    ws.cell(row=row_num, column=21).alignment = LEFT
    ws.cell(row=row_num, column=22).value = centro
    ws.cell(row=row_num, column=22).font = DATA_FONT
    ws.cell(row=row_num, column=22).alignment = LEFT

    # Multimoneda (cols 23-24)
    importe_ars = data.get('importe', 0) or 0
    tc = get_usd_rate()
    importe_usd = convertir_a_usd(importe_ars, tc)
    ws.cell(row=row_num, column=23).value = importe_usd
    ws.cell(row=row_num, column=23).font = DATA_FONT
    ws.cell(row=row_num, column=23).alignment = LEFT
    ws.cell(row=row_num, column=24).value = tc if tc else "N/D"
    ws.cell(row=row_num, column=24).font = DATA_FONT
    ws.cell(row=row_num, column=24).alignment = CENTER

    # Detección de Anomalías (col 25)
    alertas = detectar_anomalias(data)
    alertas_str = " | ".join(alertas) if alertas else "OK"
    cell_alertas = ws.cell(row=row_num, column=25)
    cell_alertas.value = alertas_str
    cell_alertas.font = RED_FONT if alertas else DATA_FONT
    cell_alertas.alignment = LEFT
    if alertas:
        cell_alertas.fill = ORANGE_FILL

    # Archivo PDF (col 26)
    cell_u = ws.cell(row=row_num, column=26)
    cell_u.value = filename
    cell_u.font = DATA_FONT
    cell_u.alignment = LEFT

    # Guardar datos extra en data dict para el response
    data['_categoria'] = categoria
    data['_centro'] = centro
    data['_usd'] = importe_usd
    data['_tc'] = tc
    data['_alertas'] = alertas

    return missing


def add_otros_doc(wb, filename, observacion):
    if "Otros Documentos" not in wb.sheetnames:
        return
    ws2 = wb["Otros Documentos"]
    existing = get_existing_otros(ws2)
    if filename in existing:
        return
    row_num = max(ws2.max_row + 1, 5)
    ext = os.path.splitext(filename)[1].upper().replace('.', '') or 'Desconocido'
    ws2.cell(row=row_num, column=1).value = filename
    ws2.cell(row=row_num, column=1).font = DATA_FONT
    ws2.cell(row=row_num, column=2).value = ext
    ws2.cell(row=row_num, column=2).font = DATA_FONT
    ws2.cell(row=row_num, column=3).value = observacion
    ws2.cell(row=row_num, column=3).font = DATA_FONT
    ws2.cell(row=row_num, column=4).value = datetime.now().strftime("%d/%m/%Y")
    ws2.cell(row=row_num, column=4).font = DATA_FONT
    ws2.cell(row=row_num, column=5).value = "Mover a otra carpeta"
    ws2.cell(row=row_num, column=5).font = DATA_FONT


# ═══════════════════════════════════════════════
# API — PROCESAMIENTO DE FACTURAS
# ═══════════════════════════════════════════════

@app.route("/ad-api/invoices/upload-process", methods=["POST"])
@requiere_auth
def upload_and_process():
    try:
        empresa_id = request.usuario["empresa_id"]
        empresa_nombre = request.usuario["empresa"]
        usuario_email = request.usuario["email"]

        files = request.files.getlist("pdfs")
        if not files or len(files) == 0:
            return jsonify({"error": True, "mensaje": "No se recibieron archivos PDF"}), 400

        wb = get_or_create_excel(empresa_id, empresa_nombre)
        ws = wb.active if wb else None
        existing = get_existing_invoices(ws) if ws else set()
        excel_path = get_empresa_excel(empresa_id)

        documents = []
        new_count = 0
        skip_count = 0
        other_count = 0

        for f in files:
            filename = f.filename or "sin_nombre.pdf"
            if not filename.lower().endswith('.pdf'):
                add_otros_doc(wb, filename, "No es un archivo PDF")
                documents.append({"nombre": filename, "estado": "no_pdf", "detalle": "No es PDF"})
                other_count += 1
                continue

            tmp_path = os.path.join(UPLOAD_DIR, filename)
            f.save(tmp_path)
            try:
                invoice_data = extract_invoice_data(tmp_path)
                if invoice_data and invoice_data.get('nro_comp') and invoice_data.get('cae'):
                    nro = invoice_data['nro_comp']
                    if nro in existing:
                        documents.append({"nombre": filename, "estado": "duplicado", "detalle": f"Comp. {nro} ya existe"})
                        skip_count += 1
                    else:
                        if ws:
                            row_num = max(ws.max_row + 1, 5)
                            missing = add_invoice_row(ws, row_num, invoice_data, filename)
                        existing.add(nro)
                        new_count += 1
                        detail = f"Comp. {nro} — ${invoice_data.get('importe', 0):,.2f}"
                        if invoice_data.get('_usd'):
                            detail += f" (USD {invoice_data['_usd']:,.2f})"
                        if missing:
                            detail += f" | Faltantes: {', '.join(missing[:3])}"
                        doc_entry = {
                            "nombre": filename, "estado": "procesado", "detalle": detail,
                            "categoria": invoice_data.get('_categoria', ''),
                            "centro": invoice_data.get('_centro', ''),
                            "alertas": invoice_data.get('_alertas', [])
                        }
                        documents.append(doc_entry)
                else:
                    add_otros_doc(wb, filename, "No es comprobante AFIP. Sin CAE ni CUIT válidos.")
                    documents.append({"nombre": filename, "estado": "no_factura", "detalle": "Sin CAE/CUIT — no es factura AFIP"})
                    other_count += 1
            except Exception as e:
                documents.append({"nombre": filename, "estado": "error", "detalle": str(e)[:60]})
                other_count += 1
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

        if wb:
            wb.save(excel_path)

        # Guardar en historial
        import json as json_lib
        conn = get_db()
        conn.execute("""INSERT INTO historial_procesamiento
            (empresa_id, usuario_email, archivos_total, facturas_nuevas, duplicadas, otros_docs, detalle)
            VALUES (?,?,?,?,?,?,?)""",
            (empresa_id, usuario_email, len(files), new_count, skip_count, other_count,
             json_lib.dumps(documents, ensure_ascii=False)))
        conn.commit()
        conn.close()

        return jsonify({
            "error": False,
            "new_invoices": new_count,
            "skipped": skip_count,
            "other_docs": other_count,
            "documents": documents,
            "excel_download": "/ad-api/invoices/download-excel",
            "mensaje": f"{new_count} facturas nuevas cargadas. {skip_count} duplicadas. {other_count} otros."
        }), 200

    except Exception as e:
        return jsonify({"error": True, "mensaje": f"Error procesando: {str(e)}"}), 500


@app.route("/ad-api/invoices/download-excel")
@requiere_auth
def download_excel():
    empresa_id = request.usuario["empresa_id"]
    excel_path = get_empresa_excel(empresa_id)
    if os.path.exists(excel_path):
        return send_from_directory(
            DATA_DIR, f"repositorio_{empresa_id}.xlsx",
            as_attachment=True, download_name="repositorio_facturas.xlsx"
        )
    return jsonify({"error": True, "mensaje": "El repositorio aún no fue creado"}), 404


@app.route("/ad-api/invoices/stats")
@requiere_auth
def invoice_stats():
    empresa_id = request.usuario["empresa_id"]
    excel_path = get_empresa_excel(empresa_id)
    if not os.path.exists(excel_path) or not EXCEL_SUPPORT:
        return jsonify({"total": 0, "exists": False})
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        total = ws.max_row - 4 if ws.max_row > 4 else 0
        wb.close()
        return jsonify({"total": total, "exists": True})
    except Exception:
        return jsonify({"total": 0, "exists": True})


@app.route("/ad-api/invoices/preview")
@requiere_auth
def invoice_preview():
    """Devuelve los datos del Excel como JSON para mostrar en la web."""
    empresa_id = request.usuario["empresa_id"]
    excel_path = get_empresa_excel(empresa_id)
    if not os.path.exists(excel_path) or not EXCEL_SUPPORT:
        return jsonify({"rows": [], "total": 0})
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row and len(row) > 2 and row[2]:
                rows.append({
                    "tipo": row[0] or "",
                    "punto_vta": row[1] or "",
                    "nro_comp": row[2] or "",
                    "fecha": row[3] or "",
                    "emisor": row[5] or "" if len(row) > 5 else "",
                    "cliente": row[9] or "" if len(row) > 9 else "",
                    "producto": row[16] or "" if len(row) > 16 else "",
                    "importe": row[17] or 0 if len(row) > 17 else 0,
                    "cae": row[18] or "" if len(row) > 18 else "",
                    "categoria": row[20] or "" if len(row) > 20 else "",
                    "centro": row[21] or "" if len(row) > 21 else "",
                    "usd": row[22] or 0 if len(row) > 22 else 0,
                    "tc": row[23] or "" if len(row) > 23 else "",
                    "alertas": row[24] or "OK" if len(row) > 24 else "OK",
                    "archivo": row[25] or "" if len(row) > 25 else (row[20] or "" if len(row) > 20 else "")
                })
        wb.close()
        return jsonify({"rows": rows, "total": len(rows)})
    except Exception as e:
        return jsonify({"rows": [], "total": 0, "error": str(e)})


@app.route("/ad-api/invoices/history")
@requiere_auth
def processing_history():
    """Devuelve el historial de procesamiento de la empresa."""
    empresa_id = request.usuario["empresa_id"]
    conn = get_db()
    rows = conn.execute("""SELECT id, usuario_email, fecha, archivos_total,
        facturas_nuevas, duplicadas, otros_docs
        FROM historial_procesamiento WHERE empresa_id=? ORDER BY id DESC LIMIT 50""",
        (empresa_id,)).fetchall()
    conn.close()
    return jsonify({"historial": [dict(r) for r in rows]})


# ═══════════════════════════════════════════════
# ADMIN — GESTIÓN DE CLIENTES / USUARIOS
# ═══════════════════════════════════════════════

def requiere_admin(f):
    @wraps(f)
    def decorado(*args, **kwargs):
        token = request.headers.get("X-AutoData-Token") or request.args.get("token")
        usuario = verificar_token(token)
        if not usuario:
            return jsonify({"error": True, "mensaje": "Sesión inválida o expirada."}), 401
        if usuario.get("rol") != "admin":
            return jsonify({"error": True, "mensaje": "Se requieren permisos de administrador."}), 403
        request.usuario = usuario
        return f(*args, **kwargs)
    return decorado


@app.route("/ad-api/admin/users", methods=["GET"])
@requiere_admin
def admin_list_users():
    """Lista todos los usuarios del sistema."""
    conn = get_db()
    rows = conn.execute("""SELECT id, email, nombre, rol, empresa, empresa_id, activo,
        ultimo_login, creado FROM usuarios ORDER BY creado DESC""").fetchall()
    conn.close()
    return jsonify({"users": [dict(r) for r in rows]})


@app.route("/ad-api/admin/users", methods=["POST"])
@requiere_admin
def admin_create_user():
    """Crea un nuevo usuario/cliente."""
    data = request.get_json() or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "").strip()
    nombre = data.get("nombre", "").strip()
    rol = data.get("rol", "operator").strip()
    empresa = data.get("empresa", "").strip()
    empresa_id = data.get("empresa_id", "").strip()

    if not email or not password or not nombre or not empresa:
        return jsonify({"error": True, "mensaje": "Faltan campos: email, password, nombre, empresa."}), 400

    if len(password) < 6:
        return jsonify({"error": True, "mensaje": "La contraseña debe tener al menos 6 caracteres."}), 400

    if rol not in ("admin", "operator"):
        rol = "operator"

    # Auto-generar empresa_id si no se provee
    if not empresa_id:
        empresa_id = re.sub(r'[^a-z0-9]', '', empresa.lower())[:20] or "empresa"

    conn = get_db()
    exists = conn.execute("SELECT id FROM usuarios WHERE email=?", (email,)).fetchone()
    if exists:
        conn.close()
        return jsonify({"error": True, "mensaje": f"Ya existe un usuario con el email {email}."}), 409

    conn.execute("""INSERT INTO usuarios (email, password_hash, nombre, rol, empresa, empresa_id)
        VALUES (?, ?, ?, ?, ?, ?)""",
        (email, _hash_password(password), nombre, rol, empresa, empresa_id))
    conn.commit()
    conn.close()

    return jsonify({
        "error": False,
        "mensaje": f"Usuario {nombre} ({email}) creado exitosamente.",
        "usuario": {"email": email, "nombre": nombre, "rol": rol, "empresa": empresa, "empresa_id": empresa_id}
    }), 201


@app.route("/ad-api/admin/users/<int:user_id>", methods=["PUT"])
@requiere_admin
def admin_update_user(user_id):
    """Actualiza un usuario existente (activar/desactivar, cambiar rol, reset password)."""
    data = request.get_json() or {}
    conn = get_db()
    user = conn.execute("SELECT * FROM usuarios WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": True, "mensaje": "Usuario no encontrado."}), 404

    updates = []
    params = []

    if "activo" in data:
        updates.append("activo=?")
        params.append(1 if data["activo"] else 0)
    if "rol" in data and data["rol"] in ("admin", "operator"):
        updates.append("rol=?")
        params.append(data["rol"])
    if "nombre" in data and data["nombre"].strip():
        updates.append("nombre=?")
        params.append(data["nombre"].strip())
    if "password" in data and len(data["password"]) >= 6:
        updates.append("password_hash=?")
        params.append(_hash_password(data["password"]))
    if "empresa" in data and data["empresa"].strip():
        updates.append("empresa=?")
        params.append(data["empresa"].strip())
    if "empresa_id" in data and data["empresa_id"].strip():
        updates.append("empresa_id=?")
        params.append(data["empresa_id"].strip())

    if not updates:
        conn.close()
        return jsonify({"error": True, "mensaje": "No se proporcionaron campos para actualizar."}), 400

    params.append(user_id)
    conn.execute(f"UPDATE usuarios SET {', '.join(updates)} WHERE id=?", params)
    conn.commit()
    conn.close()

    return jsonify({"error": False, "mensaje": "Usuario actualizado exitosamente."})


@app.route("/ad-api/admin/users/<int:user_id>", methods=["DELETE"])
@requiere_admin
def admin_delete_user(user_id):
    """Desactiva un usuario (no lo borra)."""
    conn = get_db()
    user = conn.execute("SELECT * FROM usuarios WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": True, "mensaje": "Usuario no encontrado."}), 404

    # No permitir eliminarse a sí mismo
    if user["id"] == request.usuario["id"]:
        conn.close()
        return jsonify({"error": True, "mensaje": "No podés desactivar tu propia cuenta."}), 400

    conn.execute("UPDATE usuarios SET activo=0 WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({"error": False, "mensaje": f"Usuario {user['email']} desactivado."})


# ═══════════════════════════════════════════════
# ERROR HANDLERS
# ═══════════════════════════════════════════════

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": True, "mensaje": "Recurso no encontrado."}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": True, "mensaje": "Error interno del servidor."}), 500


# ═══════════════════════════════════════════════
# INICIO — init_db se llama siempre (gunicorn + dev)
# ═══════════════════════════════════════════════

init_db()

if __name__ == "__main__":
    print(f"\n🚀 AutoData v{VERSION} — Servidor iniciado")
    print(f"   Puerto: {PORT}")
    app.run(host="0.0.0.0", port=PORT, debug=False)
